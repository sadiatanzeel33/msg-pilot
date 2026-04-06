"""Excel file parsing and export utilities."""

import io
from typing import List
from openpyxl import Workbook, load_workbook
from app.schemas.contact import ContactBulkPreview
from app.utils.phone import validate_phone


def parse_contacts_excel(file_bytes: bytes) -> List[ContactBulkPreview]:
    """Parse uploaded .xlsx and return preview rows with validation."""
    wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        return []

    # Find column indices (case-insensitive)
    header = [str(c).strip().lower() if c else "" for c in rows[0]]
    name_idx = next((i for i, h in enumerate(header) if h in ("name", "full_name", "fullname")), None)
    phone_idx = next((i for i, h in enumerate(header) if h in ("phone", "phonenumber", "phone_number", "mobile", "number")), None)
    msg_idx = next((i for i, h in enumerate(header) if h in ("message", "msg", "custom_message")), None)

    if name_idx is None or phone_idx is None:
        return [ContactBulkPreview(row=1, name="", phone="", valid=False, error="Missing required columns: Name, PhoneNumber")]

    previews: List[ContactBulkPreview] = []
    for idx, row in enumerate(rows[1:], start=2):
        name = str(row[name_idx]).strip() if row[name_idx] else ""
        phone_raw = str(row[phone_idx]).strip() if row[phone_idx] else ""
        message = str(row[msg_idx]).strip() if msg_idx is not None and row[msg_idx] else None

        if not name or not phone_raw:
            previews.append(ContactBulkPreview(row=idx, name=name, phone=phone_raw, message=message, valid=False, error="Name or phone is empty"))
            continue

        valid, normalized, err = validate_phone(phone_raw)
        previews.append(ContactBulkPreview(
            row=idx, name=name, phone=normalized, message=message,
            valid=valid, error=err if not valid else None,
        ))

    wb.close()
    return previews


def export_contacts_excel(contacts: list[dict]) -> bytes:
    """Export contacts list to xlsx bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Contacts"
    ws.append(["Name", "Phone", "Group", "Tags"])
    for c in contacts:
        ws.append([c.get("name"), c.get("phone"), c.get("group_name", ""), c.get("tags", "")])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_campaign_report(rows: list[dict]) -> bytes:
    """Export campaign delivery report to xlsx bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Campaign Report"
    ws.append(["Name", "Phone", "Status", "Sent At", "Error"])
    for r in rows:
        ws.append([r.get("name"), r.get("phone"), r.get("status"), r.get("sent_at", ""), r.get("error_message", "")])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
